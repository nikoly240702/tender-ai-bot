"""Импорт прайс-листа СИЗ команды в own_products.

Парсит Excel ('Продукция под тендеры.xlsx') → нормализует цены/единицы →
заливает в БД для category='siz'.

Использование:
    python -m scripts.import_siz_catalogue --company-id 1 --file /path/to/file.xlsx [--dry-run]

Идемпотентно: при повторном запуске удаляет старые записи category='siz'
для этой company и заливает заново.
"""
import argparse
import asyncio
import logging
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional, Tuple

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl
from sqlalchemy import delete

from database import DatabaseSession, OwnProduct, Company

logger = logging.getLogger(__name__)


def _parse_price(price_text: str) -> Tuple[Optional[Decimal], Optional[str]]:
    """Из строки '200 руб/упак 50 пар' → (Decimal('200'), 'упак 50 пар').
    Возвращает (None, None) если не разобрали."""
    if not price_text:
        return None, None
    text = str(price_text).strip()
    # Pattern: <number><spaces or comma><руб/...>
    m = re.match(r'^([\d.,\s]+?)\s*руб[/\s]*(.*)$', text, flags=re.IGNORECASE)
    if not m:
        return None, None
    num_raw = m.group(1).replace(' ', '').replace(',', '.')
    unit = (m.group(2) or '').strip().rstrip('.,;')
    try:
        return Decimal(num_raw), unit or 'шт'
    except InvalidOperation:
        return None, None


def _is_section_header(name: str) -> bool:
    """Строки типа 'ИМПОРТ (...)' / 'Производство РФ' / 'С чем можно...' — заголовки секций."""
    if not name:
        return True
    n = str(name).strip().lower()
    if not n:
        return True
    triggers = ['импорт', 'производство рф', 'с чем можно', 'дополнительно']
    return any(t in n for t in triggers)


def _parse_row(row: tuple, current_source: str) -> Optional[dict]:
    """Превращает row из Excel в dict для OwnProduct, или None если skip."""
    if not row or len(row) < 7:
        return None
    num, name, sizes, params, pack, _box, price_text = row[:7]

    if not name:
        return None
    name = str(name).strip()
    if _is_section_header(name):
        return None

    # № должен быть числом (или пустым). Если нет числа — это либо заголовок секции, либо мусор.
    try:
        int(num) if num is not None else 0
    except (ValueError, TypeError):
        return None

    price_val, price_unit = _parse_price(price_text or '')

    return {
        'name': name,
        'sizes': str(sizes).strip() if sizes else None,
        'params': str(params).strip() if params else None,
        'pack': str(pack).strip() if pack else None,
        'price': price_val,
        'price_unit': price_unit,
        'price_text': str(price_text).strip() if price_text else None,
        'source': current_source,
    }


def parse_excel(file_path: str) -> list:
    """Возвращает список dict-ов готовых к INSERT в own_products."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    items = []
    current_source = None
    for row_idx, row in enumerate(rows[1:], start=2):  # skip header
        if not row or not any(c for c in row):
            continue
        first = (row[0] or '') if row else ''
        # Если row[0] это текст и похоже на заголовок секции — обновляем source
        if isinstance(first, str) and _is_section_header(first):
            text = str(first).lower()
            if 'импорт' in text:
                current_source = 'ИМПОРТ'
            elif 'производство рф' in text or 'рф' in text:
                current_source = 'РФ'
            elif 'дополнительно' in text or 'можно учавствовать' in text:
                current_source = 'ПОД ПРОСЧЁТ'
            continue

        parsed = _parse_row(row, current_source or 'неизвестно')
        if parsed:
            items.append(parsed)

    return items


async def import_to_db(company_id: int, items: list, dry_run: bool = False) -> int:
    """Удаляет старые own_products(category='siz') этой company и заливает новые.
    Возвращает кол-во вставленных."""
    if dry_run:
        return len(items)

    async with DatabaseSession() as session:
        company = await session.get(Company, company_id)
        if not company:
            raise SystemExit(f'Company {company_id} не найдена')

        await session.execute(
            delete(OwnProduct).where(
                OwnProduct.company_id == company_id,
                OwnProduct.category == 'siz',
            )
        )
        for it in items:
            session.add(OwnProduct(
                company_id=company_id,
                category='siz',
                **it,
            ))
        await session.commit()
        return len(items)


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--company-id', type=int, required=True)
    p.add_argument('--file', type=str, required=True)
    p.add_argument('--dry-run', action='store_true')
    args = p.parse_args()

    logging.basicConfig(level=logging.INFO,
                        format='%(asctime)s %(levelname)s %(message)s')

    items = parse_excel(args.file)
    print(f'Parsed {len(items)} products from {args.file}')
    for it in items[:5]:
        print(f'  - {it["name"][:60]} → {it["price"]} {it["price_unit"]}  [{it["source"]}]')

    if not items:
        print('Nothing to import.')
        sys.exit(1)

    if args.dry_run:
        print('--dry-run: not writing to DB')
        sys.exit(0)

    inserted = asyncio.run(import_to_db(args.company_id, items))
    print(f'Inserted {inserted} rows into own_products (company_id={args.company_id}).')


if __name__ == '__main__':
    main()
