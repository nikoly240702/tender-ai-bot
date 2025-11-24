"""
Обработчики процесса поиска тендеров.
Реализует пошаговый ввод параметров:
1. Поисковый запрос
2. Ценовой диапазон
3. Количество тендеров
"""

import sys
from pathlib import Path

# Добавляем путь к корневой директории
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.fsm.context import FSMContext
from aiogram.exceptions import TelegramNetworkError, TelegramRetryAfter
import asyncio

from bot.keyboards import (
    get_tender_type_keyboard,
    get_price_range_keyboard,
    get_tender_count_keyboard,
    get_cancel_keyboard,
    get_inline_cancel_keyboard,
    get_main_menu_keyboard,
    get_tenders_list_keyboard,
    get_tender_actions_keyboard,
    get_region_keyboard,
    get_region_type_keyboard,
    get_federal_districts_keyboard,
    FEDERAL_DISTRICTS
)
from bot.states import SearchStates
from bot.config import BotConfig
from bot.db import get_database
from typing import Union
import logging

# Импортируем систему поиска
from integrated_tender_system import IntegratedTenderSystem

router = Router()
logger = logging.getLogger(__name__)

# Глобальный экземпляр системы поиска (инициализируется при первом использовании)
_tender_system = None

def get_tender_system() -> IntegratedTenderSystem:
    """Получить или создать экземпляр системы поиска."""
    global _tender_system
    if _tender_system is None:
        _tender_system = IntegratedTenderSystem()
    return _tender_system


async def safe_answer(message: Message, text: str, **kwargs) -> bool:
    """
    Безопасная отправка сообщения с retry логикой при сетевых ошибках.

    Returns:
        True если сообщение отправлено успешно, False если все попытки провалились
    """
    max_retries = 3
    retry_delay = 2  # секунды

    for attempt in range(max_retries):
        try:
            await message.answer(text, **kwargs)
            return True
        except (TelegramNetworkError, Exception) as e:
            error_name = type(e).__name__
            if attempt < max_retries - 1:
                logger.warning(
                    f"Попытка {attempt + 1}/{max_retries} отправки сообщения не удалась "
                    f"({error_name}: {str(e)}). Повтор через {retry_delay} сек..."
                )
                await asyncio.sleep(retry_delay)
                retry_delay *= 2  # Экспоненциальная задержка
            else:
                logger.error(
                    f"Не удалось отправить сообщение после {max_retries} попыток: "
                    f"{error_name}: {str(e)}"
                )
                return False
    return False


async def execute_search(
    message_or_callback: Union[Message, CallbackQuery],
    state: FSMContext,
    count: int
):
    """
    Выполняет поиск тендеров с заданными параметрами.
    Общая функция для обработки как стандартного выбора, так и кастомного ввода.

    Args:
        message_or_callback: Message или CallbackQuery объект
        state: FSM контекст
        count: Количество тендеров для поиска
    """
    # Определяем тип объекта и получаем соответствующий message
    if isinstance(message_or_callback, CallbackQuery):
        message = message_or_callback.message
        user = message_or_callback.from_user
    else:
        message = message_or_callback
        user = message.from_user

    # Сохраняем количество
    await state.update_data(max_tenders=count)

    # Получаем все параметры поиска
    data = await state.get_data()
    query = data.get('query', '')
    tender_type = data.get('tender_type')  # Тип закупки (товары/услуги/работы/None)
    price_min = data.get('price_min', 0)
    price_max = data.get('price_max', 50000000)
    regions = data.get('regions')  # Это теперь список или None

    # Форматируем для отображения
    price_min_str = f"{price_min:,}".replace(",", " ")
    price_max_str = f"{price_max:,}".replace(",", " ")

    # Форматируем регионы для отображения
    if regions is None:
        region_text = "Все регионы"
    elif len(regions) == 0:
        region_text = "Все регионы"
    elif len(regions) == 1:
        region_text = regions[0]
    elif len(regions) == 2:
        region_text = f"{regions[0]}, {regions[1]}"
    else:
        region_text = f"{regions[0]}, {regions[1]} и еще {len(regions) - 2}"

    # Форматируем тип закупки для отображения
    type_display = {
        "товары": "📦 Товары (поставка)",
        "услуги": "🔧 Услуги (обслуживание)",
        "работы": "🏗️ Работы (строительство/монтаж)",
        None: "🔍 Все типы"
    }
    type_text = type_display.get(tender_type, "🔍 Все типы")

    # Показываем сводку параметров
    params_text = (
        "📋 <b>Параметры поиска:</b>\n\n"
        f"🔍 Запрос: <b>{query}</b>\n"
        f"🎯 Тип: <b>{type_text}</b>\n"
        f"💰 Цена: <b>{price_min_str} - {price_max_str} ₽</b>\n"
        f"📍 Регион: <b>{region_text}</b>\n"
        f"🔢 Количество: <b>{count} тендеров</b>\n\n"
        f"🚀 <b>Начинаю поиск...</b>\n\n"
        f"<i>Это может занять некоторое время, пожалуйста подождите...</i>"
    )

    # Отправляем сообщение о начале поиска
    if isinstance(message_or_callback, CallbackQuery):
        await message.edit_text(params_text, parse_mode="HTML")
    else:
        await message.answer(params_text, parse_mode="HTML")

    try:
        logger.info(f"Начинаем поиск тендеров: query={query}, price_min={price_min}, price_max={price_max}, count={count}, regions={regions}")

        # Запускаем поиск в отдельном потоке (чтобы не блокировать бот)
        system = get_tender_system()

        # Оборачиваем синхронный вызов в асинхронный
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(
            None,
            lambda: system.search_and_analyze(
                search_query=query,
                price_min=price_min,
                price_max=price_max,
                max_tenders=count,
                regions=regions,  # Передаем регион
                analyze_documents=False,  # Пока не анализируем
                download_documents=False,  # Пока не скачиваем
                tender_type=tender_type  # Передаем тип закупки
            )
        )

        logger.info(f"Поиск завершен успешно, найдено тендеров: {result.get('tenders_found', 0)}")

        # Сохраняем результаты в состояние
        await state.update_data(search_results=result)
        await state.set_state(SearchStates.viewing_results)

        # Формируем список найденных тендеров
        tenders_found = result.get('tenders_found', 0)

        if tenders_found == 0:
            await message.answer(
                "😔 <b>Тендеры не найдены</b>\n\n"
                "Попробуйте изменить параметры поиска:\n"
                "• Используйте более общие термины\n"
                "• Расширьте ценовой диапазон\n"
                "• Проверьте правильность написания",
                reply_markup=get_main_menu_keyboard(),
                parse_mode="HTML"
            )
            await state.clear()
            return

        # Показываем результаты
        search_params = result.get('search_params', {})
        requested_count = search_params.get('requested_count', count)

        # Проверяем, нашли ли мы запрошенное количество
        if tenders_found < requested_count:
            results_text = (
                f"⚠️ <b>Запрошено: {requested_count} тендеров</b>\n"
                f"✅ <b>Найдено: {tenders_found} тендеров</b>\n\n"
                f"<i>К сожалению, по указанным критериям найдено меньше тендеров, "
                f"чем запрошено. Показываем все найденные результаты:</i>\n\n"
            )
        else:
            results_text = f"✅ <b>Найдено тендеров: {tenders_found}</b>\n\n"

        for i, tender_data in enumerate(result['results'][:tenders_found], 1):
            tender = tender_data['tender_info']
            number = tender.get('number', 'N/A')
            name = tender.get('name', 'Без названия')
            price = tender.get('price_formatted', 'N/A')

            # Обрезаем название, если слишком длинное
            if len(name) > 100:
                name = name[:97] + "..."

            results_text += f"{i}. <b>№ {number}</b>\n"
            results_text += f"   <b>📦 Объект закупки:</b> {name}\n"
            results_text += f"   💰 {price}\n\n"

        results_text += "<i>💡 Выберите тендер для просмотра деталей:</i>"

        # Сохраняем поиск в базу данных
        try:
            db = await get_database()

            # Обновляем информацию о пользователе
            await db.add_or_update_user(
                user_id=user.id,
                username=user.username,
                first_name=user.first_name,
                last_name=user.last_name
            )

            # Сохраняем поиск
            search_id = await db.save_search(
                user_id=user.id,
                query=query,
                price_min=price_min,
                price_max=price_max,
                tender_count=count,
                result_count=tenders_found,
                search_data=result  # Сохраняем полные результаты
            )

            # Сохраняем ID поиска в состоянии (для возможности повтора)
            await state.update_data(last_search_id=search_id)

        except Exception as e:
            # Логируем ошибку, но не прерываем работу бота
            logger.error(f"Ошибка сохранения поиска в БД: {e}")

        # Отправляем результаты с retry логикой
        success = await safe_answer(
            message,
            results_text,
            reply_markup=get_tenders_list_keyboard(tenders_found),
            parse_mode="HTML"
        )

        if not success:
            logger.error("Не удалось отправить результаты поиска пользователю")

        # Отправляем общий HTML отчет, если он был создан
        report_path = result.get('report_path')
        if report_path:
            import os
            from aiogram.types import FSInputFile

            try:
                if os.path.exists(report_path):
                    # Создаем объект файла для отправки
                    document = FSInputFile(report_path)

                    # Отправляем файл пользователю
                    await message.answer_document(
                        document=document,
                        caption=f"📊 <b>Общий отчет по поиску</b>\n\n"
                                f"🔍 Запрос: <b>{query}</b>\n"
                                f"📋 Найдено тендеров: <b>{tenders_found}</b>\n\n"
                                f"<i>Откройте файл в браузере для просмотра подробной информации по всем тендерам</i>",
                        parse_mode="HTML"
                    )
                    logger.info(f"Общий HTML отчет отправлен пользователю: {report_path}")
                else:
                    logger.warning(f"Файл общего отчета не найден: {report_path}")
            except Exception as e:
                logger.error(f"Ошибка при отправке общего HTML отчета: {e}", exc_info=True)

    except Exception as e:
        logger.error(f"Ошибка при поиске: {e}", exc_info=True)

        # Пытаемся отправить сообщение об ошибке с retry
        error_text = (
            f"❌ <b>Ошибка при поиске:</b>\n\n"
            f"<code>{str(e)}</code>\n\n"
            f"Попробуйте еще раз или обратитесь к администратору."
        )

        success = await safe_answer(
            message,
            error_text,
            reply_markup=get_main_menu_keyboard(),
            parse_mode="HTML"
        )

        if not success:
            # Если и это не помогло, логируем критическую ошибку
            logger.critical(
                f"Критическая ошибка: не удалось отправить сообщение об ошибке "
                f"пользователю {message.from_user.id}"
            )

        await state.clear()


@router.message(SearchStates.waiting_for_query)
async def process_search_query(message: Message, state: FSMContext):
    """
    Обработка поискового запроса от пользователя.
    Сохраняет запрос и переходит к выбору ценового диапазона.
    """
    query = message.text.strip()

    if not query:
        await message.answer(
            "❌ Пожалуйста, введите непустой поисковый запрос.",
            parse_mode="HTML"
        )
        return

    # Сохраняем запрос в состояние
    await state.update_data(query=query)

    # Переходим к выбору типа закупки
    await state.set_state(SearchStates.waiting_for_tender_type)

    await message.answer(
        f"✅ Запрос принят: <b>{query}</b>\n\n"
        f"🎯 <b>Шаг 2 из 4: Тип закупки</b>\n\n"
        f"Выберите тип закупки для фильтрации результатов:",
        reply_markup=get_tender_type_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.waiting_for_tender_type, F.data.startswith("type_"))
async def process_tender_type(callback: CallbackQuery, state: FSMContext):
    """
    Обработка выбора типа закупки.
    """
    await callback.answer()

    tender_type_raw = callback.data.replace("type_", "")

    # Маппинг для сохранения в состояние
    tender_type_mapping = {
        "товары": "товары",
        "услуги": "услуги",
        "работы": "работы",
        "все": None  # None означает все типы
    }

    tender_type = tender_type_mapping.get(tender_type_raw)

    # Сохраняем тип закупки
    await state.update_data(tender_type=tender_type)

    # Текст для отображения
    type_display = {
        "товары": "📦 Товары (поставка)",
        "услуги": "🔧 Услуги (обслуживание)",
        "работы": "🏗️ Работы (строительство/монтаж)",
        None: "🔍 Все типы"
    }

    # Переходим к выбору ценового диапазона
    await state.set_state(SearchStates.waiting_for_price_range)

    await callback.message.edit_text(
        f"✅ Выбран тип: <b>{type_display[tender_type]}</b>\n\n"
        f"💰 <b>Шаг 3 из 4: Ценовой диапазон</b>\n\n"
        f"Выберите диапазон цены контракта:",
        reply_markup=get_price_range_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.waiting_for_tender_type, F.data == "back_to_query")
async def back_to_query_from_type(callback: CallbackQuery, state: FSMContext):
    """Возврат к вводу поискового запроса."""
    await callback.answer()
    await state.set_state(SearchStates.waiting_for_query)

    await callback.message.edit_text(
        "🔍 <b>Шаг 1 из 4: Поисковый запрос</b>\n\n"
        "Введите ключевые слова для поиска тендеров.\n\n"
        "<i>Например: компьютерное оборудование, офисная мебель, канцтовары</i>",
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.waiting_for_price_range, F.data == "back_to_tender_type")
async def back_to_tender_type(callback: CallbackQuery, state: FSMContext):
    """Возврат к выбору типа закупки."""
    await callback.answer()
    await state.set_state(SearchStates.waiting_for_tender_type)

    data = await state.get_data()
    query = data.get('query', '')

    await callback.message.edit_text(
        f"✅ Запрос: <b>{query}</b>\n\n"
        f"🎯 <b>Шаг 2 из 4: Тип закупки</b>\n\n"
        f"Выберите тип закупки для фильтрации результатов:",
        reply_markup=get_tender_type_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.waiting_for_price_range, F.data == "cancel")
async def cancel_price_range(callback: CallbackQuery, state: FSMContext):
    """Отмена при выборе ценового диапазона."""
    await cancel_action(callback, state)


@router.callback_query(SearchStates.waiting_for_price_range, F.data == "back_to_query")
async def back_to_query(callback: CallbackQuery, state: FSMContext):
    """Возврат к вводу поискового запроса."""
    await callback.answer()

    # Возвращаемся к состоянию ввода запроса
    await state.set_state(SearchStates.waiting_for_query)

    await callback.message.edit_text(
        "🔍 <b>Шаг 1 из 3: Поисковый запрос</b>\n\n"
        "Введите ваш запрос для поиска тендеров.\n\n"
        "Примеры:\n"
        "• поставка компьютеров\n"
        "• разработка программного обеспечения\n"
        "• ремонт дорог\n\n"
        "<i>Или нажмите \"Отмена\" для возврата в главное меню</i>",
        parse_mode="HTML",
        reply_markup=get_inline_cancel_keyboard()
    )


@router.callback_query(SearchStates.waiting_for_price_range, F.data.startswith("price_"))
async def process_price_range(callback: CallbackQuery, state: FSMContext):
    """
    Обработка выбора ценового диапазона.
    Либо применяет preset, либо запрашивает custom ввод.
    """
    await callback.answer()

    price_option = callback.data.replace("price_", "")

    if price_option == "custom":
        # Запрашиваем минимальную цену
        await state.set_state(SearchStates.waiting_for_price_min)
        await callback.message.edit_text(
            "💰 <b>Ввод своего диапазона</b>\n\n"
            "Введите <b>минимальную</b> цену контракта в рублях:\n\n"
            "Примеры:\n"
            "• 100000\n"
            "• 500000\n"
            "• 0 (без минимума)\n\n"
            "<i>Или нажмите \"Отмена\" для возврата</i>",
            parse_mode="HTML",
            reply_markup=get_inline_cancel_keyboard()
        )
    else:
        # Применяем preset
        price_range = BotConfig.PRICE_RANGES.get(price_option)
        if not price_range:
            await callback.message.answer("❌ Ошибка: неверный диапазон цен")
            return

        await state.update_data(
            price_min=price_range[0],
            price_max=price_range[1]
        )

        # Переходим к выбору региона
        await state.set_state(SearchStates.waiting_for_region)

        # Форматируем цены для отображения
        price_min_str = f"{price_range[0]:,}".replace(",", " ")
        price_max_str = f"{price_range[1]:,}".replace(",", " ")

        # Инициализируем пустой список выбранных регионов и федеральных округов
        await state.update_data(selected_regions=[], selected_districts=[])

        await callback.message.edit_text(
            f"✅ Ценовой диапазон: <b>{price_min_str} - {price_max_str} ₽</b>\n\n"
            f"📍 <b>Шаг 3 из 4: Регион</b>\n\n"
            f"Выберите способ фильтрации по географии:",
            reply_markup=get_region_type_keyboard(),
            parse_mode="HTML"
        )


@router.message(SearchStates.waiting_for_price_min)
async def process_price_min(message: Message, state: FSMContext):
    """
    Обработка ввода минимальной цены.
    """
    try:
        price_min = int(message.text.strip())
        if price_min < 0:
            await message.answer(
                "❌ Минимальная цена не может быть отрицательной.\n"
                "Попробуйте еще раз:",
                parse_mode="HTML"
            )
            return

        # Сохраняем минимальную цену
        await state.update_data(price_min=price_min)

        # Запрашиваем максимальную цену
        await state.set_state(SearchStates.waiting_for_price_max)

        price_min_str = f"{price_min:,}".replace(",", " ")
        await message.answer(
            f"✅ Минимальная цена: <b>{price_min_str} ₽</b>\n\n"
            f"Теперь введите <b>максимальную</b> цену контракта в рублях:\n\n"
            f"Примеры:\n"
            f"• 1000000\n"
            f"• 5000000\n"
            f"• 50000000 (без максимума используйте большое число)\n\n"
            f"<i>Или нажмите \"Отмена\" для возврата</i>",
            parse_mode="HTML",
            reply_markup=get_cancel_keyboard()
        )

    except ValueError:
        await message.answer(
            "❌ Пожалуйста, введите число (например: 500000).\n"
            "Попробуйте еще раз:",
            parse_mode="HTML"
        )


@router.message(SearchStates.waiting_for_price_max)
async def process_price_max(message: Message, state: FSMContext):
    """
    Обработка ввода максимальной цены.
    Переходит к выбору количества тендеров.
    """
    try:
        price_max = int(message.text.strip())
        if price_max < 0:
            await message.answer(
                "❌ Максимальная цена не может быть отрицательной.\n"
                "Попробуйте еще раз:",
                parse_mode="HTML"
            )
            return

        # Получаем минимальную цену из состояния
        data = await state.get_data()
        price_min = data.get('price_min', 0)

        if price_max < price_min:
            await message.answer(
                f"❌ Максимальная цена не может быть меньше минимальной ({price_min:,} ₽).\n"
                f"Попробуйте еще раз:",
                parse_mode="HTML"
            )
            return

        # Сохраняем максимальную цену
        await state.update_data(price_max=price_max)

        # Переходим к выбору региона
        await state.set_state(SearchStates.waiting_for_region)

        price_min_str = f"{price_min:,}".replace(",", " ")
        price_max_str = f"{price_max:,}".replace(",", " ")

        # Инициализируем пустой список выбранных регионов и федеральных округов
        await state.update_data(selected_regions=[], selected_districts=[])

        await message.answer(
            f"✅ Ценовой диапазон: <b>{price_min_str} - {price_max_str} ₽</b>\n\n"
            f"📍 <b>Шаг 3 из 4: Регион</b>\n\n"
            f"Выберите способ фильтрации по географии:",
            reply_markup=get_region_type_keyboard(),
            parse_mode="HTML"
        )

    except ValueError:
        await message.answer(
            "❌ Пожалуйста, введите число (например: 5000000).\n"
            "Попробуйте еще раз:",
            parse_mode="HTML"
        )


# Обработчики выбора типа фильтра (регионы или федеральные округа)

@router.callback_query(SearchStates.waiting_for_region, F.data == "cancel")
async def cancel_region_selection(callback: CallbackQuery, state: FSMContext):
    """Отмена при выборе региона."""
    await cancel_action(callback, state)


@router.callback_query(SearchStates.waiting_for_region, F.data == "region_type_regions")
async def choose_regions(callback: CallbackQuery, state: FSMContext):
    """Выбор фильтрации по регионам."""
    await callback.answer()

    # Инициализируем пустой список выбранных регионов
    await state.update_data(selected_regions=[])

    await callback.message.edit_text(
        f"📍 <b>Шаг 3 из 4: Регион</b>\n\n"
        f"Выберите один или несколько регионов:\n"
        f"<i>(нажмите на регион чтобы выбрать/снять)</i>",
        reply_markup=get_region_keyboard([]),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.waiting_for_region, F.data == "region_type_districts")
async def choose_districts(callback: CallbackQuery, state: FSMContext):
    """Выбор фильтрации по федеральным округам."""
    await callback.answer()

    # Инициализируем пустой список выбранных федеральных округов
    await state.update_data(selected_districts=[])

    await callback.message.edit_text(
        f"🌐 <b>Федеральные округа</b>\n\n"
        f"Выберите один или несколько федеральных округов:\n"
        f"<i>(нажмите на округ чтобы выбрать/снять)</i>",
        reply_markup=get_federal_districts_keyboard([]),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.waiting_for_region, F.data == "region_type_back")
async def back_to_region_type(callback: CallbackQuery, state: FSMContext):
    """Возврат к выбору типа фильтра."""
    await callback.answer()

    await callback.message.edit_text(
        f"📍 <b>Шаг 3 из 4: Регион</b>\n\n"
        f"Выберите способ фильтрации по географии:",
        reply_markup=get_region_type_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.waiting_for_region, F.data == "back_to_price")
async def back_to_price(callback: CallbackQuery, state: FSMContext):
    """Возврат к выбору ценового диапазона."""
    await callback.answer()

    # Получаем сохраненный запрос
    data = await state.get_data()
    query = data.get('query', 'ваш запрос')

    # Возвращаемся к состоянию выбора цены
    await state.set_state(SearchStates.waiting_for_price_range)

    # Очищаем данные о регионах
    await state.update_data(selected_regions=[], selected_districts=[])

    await callback.message.edit_text(
        f"✅ Запрос: <b>{query}</b>\n\n"
        f"💰 <b>Шаг 2 из 3: Ценовой диапазон</b>\n\n"
        f"Выберите диапазон цены контракта:",
        reply_markup=get_price_range_keyboard(),
        parse_mode="HTML"
    )


# Обработчики федеральных округов

@router.callback_query(SearchStates.waiting_for_region, F.data.startswith("district_toggle_"))
async def toggle_district(callback: CallbackQuery, state: FSMContext):
    """Переключение выбора федерального округа."""
    await callback.answer()

    district_code = callback.data.replace("district_toggle_", "")

    # Получаем текущий список выбранных округов
    data = await state.get_data()
    selected_districts = data.get('selected_districts', [])

    # Переключаем выбор
    if district_code in selected_districts:
        selected_districts.remove(district_code)
    else:
        selected_districts.append(district_code)

    # Сохраняем обновленный список
    await state.update_data(selected_districts=selected_districts)

    # Обновляем клавиатуру
    await callback.message.edit_reply_markup(
        reply_markup=get_federal_districts_keyboard(selected_districts)
    )


@router.callback_query(SearchStates.waiting_for_region, F.data == "district_clear")
async def clear_districts(callback: CallbackQuery, state: FSMContext):
    """Сброс выбора федеральных округов."""
    await callback.answer("Выбор сброшен")

    await state.update_data(selected_districts=[])

    # Обновляем клавиатуру
    await callback.message.edit_reply_markup(
        reply_markup=get_federal_districts_keyboard([])
    )


@router.callback_query(SearchStates.waiting_for_region, F.data == "district_confirm")
async def confirm_districts(callback: CallbackQuery, state: FSMContext):
    """Подтверждение выбора федеральных округов."""
    await callback.answer()

    # Получаем выбранные округа
    data = await state.get_data()
    selected_districts = data.get('selected_districts', [])

    if not selected_districts:
        await callback.answer("⚠️ Выберите хотя бы один федеральный округ", show_alert=True)
        return

    # Преобразуем федеральные округа в список регионов
    regions = []
    for district_code in selected_districts:
        district_data = FEDERAL_DISTRICTS.get(district_code)
        if district_data:
            regions.extend(district_data['regions'])

    # Сохраняем регионы
    await state.update_data(regions=regions)

    # Переходим к выбору количества тендеров
    await state.set_state(SearchStates.waiting_for_tender_count)

    # Формируем строку с выбранными округами
    districts_names = [FEDERAL_DISTRICTS[code]['name'] for code in selected_districts]
    districts_str = ", ".join(districts_names)

    await callback.message.edit_text(
        f"✅ Выбрано федеральных округов: <b>{len(selected_districts)}</b>\n"
        f"<i>{districts_str}</i>\n"
        f"(всего регионов: {len(regions)})\n\n"
        f"🔢 <b>Шаг 4 из 4: Количество тендеров</b>\n\n"
        f"Сколько тендеров найти?",
        reply_markup=get_tender_count_keyboard(),
        parse_mode="HTML"
    )


# Обработчики регионов

@router.callback_query(SearchStates.waiting_for_region, F.data.startswith("region_toggle_"))
async def toggle_region(callback: CallbackQuery, state: FSMContext):
    """Переключение выбора региона (добавление/удаление)."""
    # Сразу отвечаем на callback, чтобы убрать "часики" в Telegram
    await callback.answer()

    region_name = callback.data.replace("region_toggle_", "")

    # Получаем текущий список выбранных регионов
    data = await state.get_data()
    selected_regions = data.get('selected_regions', [])

    # Переключаем выбор
    if region_name in selected_regions:
        selected_regions.remove(region_name)
    else:
        selected_regions.append(region_name)

    # Сохраняем обновленный список
    await state.update_data(selected_regions=selected_regions)

    # Обновляем клавиатуру
    await callback.message.edit_reply_markup(
        reply_markup=get_region_keyboard(selected_regions)
    )


@router.callback_query(SearchStates.waiting_for_region, F.data == "region_clear")
async def clear_regions(callback: CallbackQuery, state: FSMContext):
    """Сброс выбора регионов."""
    # Сразу отвечаем на callback
    await callback.answer("Выбор сброшен")

    await state.update_data(selected_regions=[])

    # Обновляем клавиатуру
    await callback.message.edit_reply_markup(
        reply_markup=get_region_keyboard([])
    )


@router.callback_query(SearchStates.waiting_for_region, F.data == "region_all")
async def select_all_regions(callback: CallbackQuery, state: FSMContext):
    """Выбор всех регионов (без фильтра)."""
    await callback.answer()

    # Сохраняем пустой список (означает "все регионы")
    await state.update_data(regions=None)

    # Переходим к выбору количества тендеров
    await state.set_state(SearchStates.waiting_for_tender_count)

    await callback.message.edit_text(
        f"✅ Регион: <b>Все регионы</b>\n\n"
        f"🔢 <b>Шаг 4 из 4: Количество тендеров</b>\n\n"
        f"Сколько тендеров найти?",
        reply_markup=get_tender_count_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.waiting_for_region, F.data == "region_confirm")
async def confirm_regions(callback: CallbackQuery, state: FSMContext):
    """Подтверждение выбора регионов."""
    # Получаем выбранные регионы
    data = await state.get_data()
    selected_regions = data.get('selected_regions', [])

    if not selected_regions:
        # Если ничего не выбрано, показываем alert
        await callback.answer("⚠️ Выберите хотя бы один регион", show_alert=True)
        return

    # Если регионы выбраны, подтверждаем
    await callback.answer()

    # Сохраняем список регионов для поиска
    await state.update_data(regions=selected_regions)

    # Переходим к выбору количества тендеров
    await state.set_state(SearchStates.waiting_for_tender_count)

    # Форматируем список регионов для отображения
    if len(selected_regions) == 1:
        region_text = selected_regions[0]
    elif len(selected_regions) == 2:
        region_text = f"{selected_regions[0]}, {selected_regions[1]}"
    else:
        region_text = f"{selected_regions[0]}, {selected_regions[1]} и еще {len(selected_regions) - 2}"

    await callback.message.edit_text(
        f"✅ Регионы: <b>{region_text}</b>\n\n"
        f"🔢 <b>Шаг 4 из 4: Количество тендеров</b>\n\n"
        f"Сколько тендеров найти?",
        reply_markup=get_tender_count_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.waiting_for_region, F.data == "region_custom")
async def request_custom_region(callback: CallbackQuery, state: FSMContext):
    """Обработка кнопки 'Ввести регион вручную'."""
    await callback.answer()

    # Переходим в состояние ожидания ручного ввода региона
    await state.set_state(SearchStates.waiting_for_custom_region)

    await callback.message.edit_text(
        "✍️ <b>Ручной ввод региона</b>\n\n"
        "Введите название региона:\n\n"
        "Примеры:\n"
        "• Ростовская область\n"
        "• Республика Крым\n"
        "• Приморский край\n"
        "• Тюменская область\n\n"
        "<i>💡 Можно вводить несколько регионов через запятую</i>",
        parse_mode="HTML"
    )


@router.message(SearchStates.waiting_for_custom_region)
async def process_custom_region(message: Message, state: FSMContext):
    """Обработка ручного ввода региона."""
    region_input = message.text.strip()

    if not region_input:
        await message.answer(
            "⚠️ Пожалуйста, введите название региона",
            parse_mode="HTML"
        )
        return

    # Разбиваем по запятой, если несколько регионов
    regions = [r.strip() for r in region_input.split(',') if r.strip()]

    # Сохраняем регионы
    await state.update_data(regions=regions)

    # Переходим к выбору количества тендеров
    await state.set_state(SearchStates.waiting_for_tender_count)

    # Формируем текст с регионами
    if len(regions) == 1:
        region_text = regions[0]
    elif len(regions) == 2:
        region_text = f"{regions[0]}, {regions[1]}"
    else:
        region_text = f"{regions[0]}, {regions[1]} и еще {len(regions) - 2}"

    await message.answer(
        f"✅ Регионы: <b>{region_text}</b>\n\n"
        f"🔢 <b>Шаг 4 из 4: Количество тендеров</b>\n\n"
        f"Сколько тендеров найти?",
        reply_markup=get_tender_count_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.waiting_for_tender_count, F.data == "cancel")
async def cancel_tender_count(callback: CallbackQuery, state: FSMContext):
    """Отмена при выборе количества тендеров."""
    await cancel_action(callback, state)


@router.callback_query(SearchStates.waiting_for_tender_count, F.data == "back_to_region")
async def back_to_region(callback: CallbackQuery, state: FSMContext):
    """Возврат к выбору региона."""
    await callback.answer()

    # Получаем данные для формирования сообщения
    data = await state.get_data()
    query = data.get('query', 'ваш запрос')
    price_min = data.get('price_min', 0)
    price_max = data.get('price_max', 0)

    # Форматируем цены
    price_min_str = f"{price_min:,}".replace(",", " ")
    price_max_str = f"{price_max:,}".replace(",", " ")

    # Возвращаемся к состоянию выбора региона
    await state.set_state(SearchStates.waiting_for_region)

    await callback.message.edit_text(
        f"✅ Запрос: <b>{query}</b>\n"
        f"✅ Ценовой диапазон: <b>{price_min_str} - {price_max_str} ₽</b>\n\n"
        f"📍 <b>Шаг 3 из 4: Регион</b>\n\n"
        f"Выберите способ фильтрации по географии:",
        reply_markup=get_region_type_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.waiting_for_tender_count, F.data == "count_custom")
async def process_custom_count_request(callback: CallbackQuery, state: FSMContext):
    """
    Обработка выбора 'Свой вариант' для количества тендеров.
    """
    await callback.answer()

    # Переходим в состояние ожидания кастомного ввода
    await state.set_state(SearchStates.waiting_for_custom_count)

    # Убираем inline клавиатуру, чтобы пользователь мог вводить текст
    await callback.message.edit_text(
        "💯 <b>Свой вариант</b>\n\n"
        "Введите количество тендеров для поиска:\n\n"
        "Примеры:\n"
        "• 1\n"
        "• 7\n"
        "• 15\n"
        "• 20\n\n"
        "<i>Рекомендуется: от 3 до 10 тендеров</i>",
        reply_markup=None,  # Убираем inline клавиатуру
        parse_mode="HTML"
    )


@router.message(SearchStates.waiting_for_custom_count)
async def process_custom_count(message: Message, state: FSMContext):
    """
    Обработка ввода кастомного количества тендеров.
    """
    try:
        count = int(message.text.strip())

        if count < 1:
            await message.answer(
                "❌ Количество должно быть не менее 1.\n"
                "Попробуйте еще раз:",
                parse_mode="HTML"
            )
            return

        if count > 50:
            await message.answer(
                "⚠️ Количество слишком большое!\n\n"
                "Для оптимальной работы рекомендуется не более 50 тендеров.\n"
                "Попробуйте еще раз:",
                parse_mode="HTML"
            )
            return

        # Запускаем поиск с кастомным количеством
        await execute_search(message, state, count)

    except ValueError:
        await message.answer(
            "❌ Пожалуйста, введите число (например: 7).\n"
            "Попробуйте еще раз:",
            parse_mode="HTML"
        )


@router.callback_query(SearchStates.waiting_for_tender_count, F.data.startswith("count_"))
async def process_tender_count(callback: CallbackQuery, state: FSMContext):
    """
    Обработка выбора количества тендеров из предложенных вариантов.
    Запускает процесс поиска.
    """
    # Проверяем, что это не "count_custom"
    if callback.data == "count_custom":
        return

    await callback.answer()

    count = int(callback.data.replace("count_", ""))

    # Используем общую функцию для выполнения поиска
    await execute_search(callback, state, count)


@router.callback_query(F.data == "cancel", flags={"priority": -10})
async def cancel_action(callback: CallbackQuery, state: FSMContext):
    """
    Отмена текущего действия и возврат в главное меню.
    Работает в любом состоянии.
    """
    await callback.answer()

    # Получаем текущее состояние для логирования
    current_state = await state.get_state()
    if current_state:
        logger.info(f"Отмена действия из состояния: {current_state}")

    await state.clear()

    # Используем safe_answer для надежной отправки
    await callback.message.edit_text(
        "❌ Действие отменено.",
        parse_mode="HTML"
    )

    await safe_answer(
        callback.message,
        "Вы вернулись в главное меню.",
        reply_markup=get_main_menu_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.viewing_results, F.data.startswith("details_"))
async def show_tender_details(callback: CallbackQuery, state: FSMContext):
    """
    Показать подробную информацию о тендере.
    """
    await callback.answer()

    # Получаем индекс тендера
    tender_index = int(callback.data.replace("details_", ""))

    # Получаем результаты поиска из состояния
    data = await state.get_data()
    search_results = data.get('search_results', {})
    results = search_results.get('results', [])

    if tender_index >= len(results):
        await callback.message.answer(
            "❌ Тендер не найден",
            parse_mode="HTML"
        )
        return

    # Получаем информацию о тендере
    tender_data = results[tender_index]
    tender = tender_data['tender_info']

    # Формируем подробное описание
    details_text = "📋 <b>ПОДРОБНАЯ ИНФОРМАЦИЯ О ТЕНДЕРЕ</b>\n\n"

    # Основная информация
    details_text += f"<b>Номер:</b> {tender.get('number', 'N/A')}\n\n"
    details_text += f"<b>Название:</b>\n{tender.get('name', 'Без названия')}\n\n"
    details_text += f"<b>💰 Цена:</b> {tender.get('price_formatted', 'N/A')}\n\n"

    # Заказчик
    customer = tender.get('customer', 'Не указан')
    if len(customer) > 100:
        customer = customer[:97] + "..."
    details_text += f"<b>🏢 Заказчик:</b>\n{customer}\n\n"

    # Сроки
    if tender.get('publish_date'):
        details_text += f"<b>📅 Опубликован:</b> {tender['publish_date']}\n"
    if tender.get('deadline'):
        details_text += f"<b>⏰ Окончание подачи:</b> {tender['deadline']}\n"
    if tender.get('contract_execution_date'):
        details_text += f"<b>📆 Срок исполнения:</b> {tender['contract_execution_date']}\n"

    details_text += "\n"

    # Требования
    if tender.get('requirements'):
        req = tender['requirements']
        if len(req) > 300:
            req = req[:297] + "..."
        details_text += f"<b>📝 Требования:</b>\n{req}\n\n"

    # Ссылка на zakupki.gov.ru
    tender_url = tender.get('url', '')

    # Если URL относительный, добавляем домен
    if tender_url and not tender_url.startswith('http'):
        tender_url = f"https://zakupki.gov.ru{tender_url}"

    # Переходим в состояние просмотра деталей
    await state.set_state(SearchStates.viewing_tender_details)
    await state.update_data(current_tender_index=tender_index)

    # Отправляем сообщение с деталями и правильной клавиатурой
    await callback.message.edit_text(
        details_text,
        parse_mode="HTML",
        reply_markup=get_tender_actions_keyboard(
            tender_index,
            tender_url=tender_url,
            has_analysis=tender_data.get('analysis_success', False)
        )
    )


@router.callback_query(SearchStates.viewing_tender_details, F.data == "back_to_results")
async def back_to_results(callback: CallbackQuery, state: FSMContext):
    """
    Вернуться к списку результатов поиска.
    """
    await callback.answer()

    # Возвращаемся в состояние просмотра результатов
    await state.set_state(SearchStates.viewing_results)

    # Получаем результаты из состояния
    data = await state.get_data()
    search_results = data.get('search_results', {})
    tenders_found = search_results.get('tenders_found', 0)
    results = search_results.get('results', [])

    # Формируем список снова
    results_text = f"✅ <b>Найдено тендеров: {tenders_found}</b>\n\n"

    for i, tender_data in enumerate(results[:tenders_found], 1):
        tender = tender_data['tender_info']
        number = tender.get('number', 'N/A')
        name = tender.get('name', 'Без названия')
        price = tender.get('price_formatted', 'N/A')

        if len(name) > 100:
            name = name[:97] + "..."

        results_text += f"{i}. <b>№ {number}</b>\n"
        results_text += f"   <b>📦 Объект закупки:</b> {name}\n"
        results_text += f"   💰 {price}\n\n"

    results_text += "<i>💡 Выберите тендер для просмотра деталей:</i>"

    await callback.message.edit_text(
        results_text,
        reply_markup=get_tenders_list_keyboard(tenders_found),
        parse_mode="HTML"
    )


@router.callback_query(SearchStates.viewing_tender_details, F.data.startswith("analyze_"))
async def analyze_tender(callback: CallbackQuery, state: FSMContext):
    """
    Запустить AI-анализ документов тендера.
    """
    await callback.answer()

    # Получаем индекс тендера
    tender_index = int(callback.data.replace("analyze_", ""))

    # Получаем результаты поиска из состояния
    data = await state.get_data()
    search_results = data.get('search_results', {})
    results = search_results.get('results', [])

    if tender_index >= len(results):
        await callback.message.answer(
            "❌ Тендер не найден",
            parse_mode="HTML"
        )
        return

    tender_data = results[tender_index]
    tender = tender_data['tender_info']

    # Получаем и нормализуем URL
    tender_url = tender.get('url', '')
    if tender_url and not tender_url.startswith('http'):
        tender_url = f"https://zakupki.gov.ru{tender_url}"

    # Показываем сообщение о начале анализа
    await callback.message.edit_text(
        "🤖 <b>Запускаю AI-анализ...</b>\n\n"
        f"📄 Тендер: {tender.get('number', 'N/A')}\n\n"
        "⏳ <b>Этапы:</b>\n"
        "1️⃣ Скачивание документов...\n"
        "2️⃣ Извлечение текста из PDF\n"
        "3️⃣ AI-анализ документации\n"
        "4️⃣ Формирование отчета\n\n"
        "<i>Это может занять 1-3 минуты, пожалуйста подождите...</i>",
        parse_mode="HTML"
    )

    try:
        # Получаем систему
        system = get_tender_system()

        # Оборачиваем в асинхронный вызов
        loop = asyncio.get_event_loop()

        # Скачиваем документы
        await callback.message.edit_text(
            "🤖 <b>AI-анализ в процессе...</b>\n\n"
            f"📄 Тендер: {tender.get('number', 'N/A')}\n\n"
            "⏳ <b>Текущий этап:</b>\n"
            "1️⃣ ✅ Скачивание документов...\n"
            "2️⃣ 🔄 Извлечение текста из PDF\n"
            "3️⃣ ⏸ AI-анализ документации\n"
            "4️⃣ ⏸ Формирование отчета",
            parse_mode="HTML"
        )

        download_result = await loop.run_in_executor(
            None,
            lambda: system.document_downloader.download_documents(
                tender_url=tender_url,
                tender_number=tender.get('number', 'unknown'),
                doc_types=None
            )
        )

        if download_result['downloaded'] == 0:
            await callback.message.edit_text(
                "⚠️ <b>Не удалось скачать документы</b>\n\n"
                f"📄 Тендер: {tender.get('number', 'N/A')}\n\n"
                "Возможные причины:\n"
                "• Документы не опубликованы\n"
                "• Проблема с доступом к сайту\n"
                "• Неверный URL",
                reply_markup=get_tender_actions_keyboard(
                    tender_index,
                    tender_url=tender_url,
                    has_analysis=False
                ),
                parse_mode="HTML"
            )
            return

        # Анализируем документы
        await callback.message.edit_text(
            "🤖 <b>AI-анализ в процессе...</b>\n\n"
            f"📄 Тендер: {tender.get('number', 'N/A')}\n\n"
            "⏳ <b>Текущий этап:</b>\n"
            "1️⃣ ✅ Скачивание документов\n"
            "2️⃣ ✅ Извлечение текста из PDF\n"
            "3️⃣ 🔄 AI-анализ документации...\n"
            "4️⃣ ⏸ Формирование отчета\n\n"
            f"<i>Анализирую {download_result['downloaded']} документ(ов)...</i>",
            parse_mode="HTML"
        )

        # Получаем пути к файлам
        file_paths = [doc['path'] for doc in download_result.get('files', [])]

        # Импортируем агента для анализа
        from main import TenderAnalysisAgent
        from bot.db import get_database

        # Создаем агента и анализируем
        agent = TenderAnalysisAgent()

        # Инициализируем БД для кэширования
        agent.db = await get_database()

        # Анализируем с кэшированием (теперь метод async)
        tender_num = tender.get('number', 'unknown')
        analysis_result = await agent.analyze_tender(
            file_paths,
            tender_number=tender_num,
            use_cache=True
        )

        # Проверяем, что анализ вернул результаты
        if not analysis_result or not isinstance(analysis_result, dict):
            await callback.message.edit_text(
                "⚠️ <b>Анализ не удался</b>\n\n"
                f"📄 Тендер: {tender.get('number', 'N/A')}\n\n"
                "Возможные причины:\n"
                "• Документы в формате ZIP (не поддерживается)\n"
                "• Не удалось извлечь текст из PDF\n"
                "• Документы защищены паролем\n\n"
                "<i>Попробуйте другой тендер или скачайте документы вручную</i>",
                reply_markup=get_tender_actions_keyboard(
                    tender_index,
                    tender_url=tender_url,
                    has_analysis=False
                ),
                parse_mode="HTML"
            )
            return

        # Получаем пути к отчетам
        report_paths = analysis_result.get('report_paths', {})
        logger.info(f"Report paths получены из analysis_result: {report_paths}")
        html_path = report_paths.get('html') if report_paths else None
        logger.info(f"HTML path извлечен: {html_path}")

        # Проверяем наличие ошибок при генерации отчета
        report_error = analysis_result.get('report_generation_error')
        if report_error:
            logger.error(f"Ошибка генерации отчета: {report_error}")

        # Обновляем данные в состоянии
        results[tender_index]['documents_downloaded'] = download_result.get('files', [])
        results[tender_index]['download_success'] = True
        results[tender_index]['analysis_result'] = analysis_result
        results[tender_index]['analysis_success'] = True
        results[tender_index]['tender_dir'] = download_result.get('tender_dir')
        results[tender_index]['html_report_path'] = html_path

        search_results['results'] = results
        await state.update_data(search_results=search_results)

        # Отправляем HTML отчет пользователю
        html_sent = False
        logger.info(f"Проверка HTML отчета: html_path = {html_path}")
        if html_path:
            try:
                import os
                from aiogram.types import FSInputFile

                logger.info(f"HTML путь задан: {html_path}")
                if os.path.exists(html_path):
                    logger.info(f"HTML файл существует, отправляем пользователю...")
                    # Создаем объект файла для отправки
                    document = FSInputFile(html_path)

                    # Отправляем файл пользователю
                    await callback.message.answer_document(
                        document=document,
                        caption=f"📊 <b>Подробный AI-анализ тендера {tender.get('number', 'N/A')}</b>\n\nОткройте файл в браузере для просмотра полного анализа",
                        parse_mode="HTML"
                    )
                    html_sent = True
                    logger.info(f"✅ HTML отчет успешно отправлен пользователю: {html_path}")
                else:
                    logger.warning(f"⚠️ HTML файл не найден по пути: {html_path}")
            except Exception as e:
                logger.error(f"❌ Не удалось отправить HTML отчет: {e}", exc_info=True)
        else:
            logger.warning("⚠️ HTML путь не задан (html_path is None or empty)")

        # Формируем сообщение с результатами
        results_text = "✅ <b>AI-АНАЛИЗ ЗАВЕРШЕН</b>\n\n"
        results_text += f"📄 <b>Тендер:</b> {tender.get('number', 'N/A')}\n"
        results_text += f"📥 <b>Документов:</b> {download_result['downloaded']}\n"

        # Добавляем информацию о HTML отчете
        if html_sent:
            results_text += "📊 <b>HTML отчет:</b> отправлен\n"
        elif report_error:
            results_text += f"⚠️ <b>HTML отчет:</b> ошибка генерации\n"
        else:
            results_text += "⚠️ <b>HTML отчет:</b> не создан\n"

        # Получаем summary - может быть на разных уровнях вложенности
        summary = analysis_result.get('analysis_summary') if analysis_result else {}
        if not summary:
            summary = analysis_result.get('summary') if analysis_result else {}
        if not summary:
            summary = {}

        # Проверяем разные возможные структуры с защитой от None
        tender_info = analysis_result.get('tender_info') if analysis_result else None
        if not tender_info or not isinstance(tender_info, dict):
            tender_info = {}

        requirements = analysis_result.get('requirements') if analysis_result else None
        if not requirements or not isinstance(requirements, dict):
            requirements = {}

        gaps = analysis_result.get('gaps') if analysis_result else None
        if not gaps or not isinstance(gaps, list):
            gaps = []

        questions_data = analysis_result.get('questions') if analysis_result else None
        if not questions_data:
            questions_data = {}

        # Проверяем, есть ли вообще данные для отображения
        has_data = bool(summary or tender_info or gaps or questions_data)

        results_text += "\n"

        # Объект закупки
        tender_name = tender_info.get('name', '')
        if tender_name and tender_name != 'N/A':
            if len(tender_name) > 150:
                tender_name = tender_name[:147] + "..."
            results_text += f"<b>📦 Объект закупки:</b>\n{tender_name}\n\n"

        # Финансы
        nmck = tender_info.get('nmck', 0)
        prepayment = tender_info.get('prepayment_percent', 0)

        if nmck or prepayment:
            results_text += "<b>💰 ФИНАНСЫ</b>\n"

            if nmck:
                results_text += f"├─ НМЦК: {nmck:,.0f} руб.\n"

            if prepayment:
                prepayment_amount = nmck * prepayment / 100 if nmck else 0
                results_text += f"├─ Аванс: {prepayment}% ({prepayment_amount:,.0f} руб.)\n"

            # Получаем данные об обеспечении из финансов
            financial = analysis_result.get('financial_analysis') if analysis_result else None
            if financial and isinstance(financial, dict):
                guarantees = financial.get('guarantees', {})
                app_guarantee = guarantees.get('application_guarantee', 0)
                contract_guarantee = guarantees.get('contract_guarantee', 0)

                if app_guarantee:
                    results_text += f"├─ Обеспечение заявки: {app_guarantee:,.0f} руб.\n"
                if contract_guarantee:
                    results_text += f"└─ Обеспечение контракта: {contract_guarantee:,.0f} руб.\n"
            else:
                # Если нет данных из financial, закрываем последнюю строку
                results_text = results_text.replace("├─ Аванс:", "└─ Аванс:")

            results_text += "\n"

        # Риски и пробелы
        if gaps and len(gaps) > 0:
            # Считаем пробелы по уровню критичности
            critical_count = sum(1 for g in gaps if isinstance(g, dict) and g.get('severity') == 'CRITICAL')
            high_count = sum(1 for g in gaps if isinstance(g, dict) and g.get('severity') == 'HIGH')
            medium_count = sum(1 for g in gaps if isinstance(g, dict) and g.get('severity') == 'MEDIUM')
            low_count = sum(1 for g in gaps if isinstance(g, dict) and g.get('severity') == 'LOW')

            results_text += "<b>⚠️ РИСКИ И ПРОБЕЛЫ</b>\n"
            severity_lines = []
            if critical_count > 0:
                severity_lines.append(f"Критические: {critical_count}")
            if high_count > 0:
                severity_lines.append(f"Важные: {high_count}")
            if medium_count > 0:
                severity_lines.append(f"Средние: {medium_count}")
            if low_count > 0:
                severity_lines.append(f"Низкие: {low_count}")

            # Форматируем с правильными префиксами
            for i, line in enumerate(severity_lines):
                if i == len(severity_lines) - 1:
                    results_text += f"└─ {line}\n"
                else:
                    results_text += f"├─ {line}\n"

            results_text += "\n"

        # Сроки
        deadline_submission = tender_info.get('deadline_submission', '')
        contract_duration = tender_info.get('contract_duration', '')

        if deadline_submission or contract_duration:
            results_text += "<b>📅 СРОКИ</b>\n"

            if deadline_submission and deadline_submission != 'N/A':
                if contract_duration and contract_duration != 'N/A':
                    results_text += f"├─ Подача заявок: до {deadline_submission}\n"
                    results_text += f"└─ Срок контракта: {contract_duration}\n"
                else:
                    results_text += f"└─ Подача заявок: до {deadline_submission}\n"
            elif contract_duration and contract_duration != 'N/A':
                results_text += f"└─ Срок контракта: {contract_duration}\n"

            results_text += "\n"

        # Краткое резюме (опционально, в конце)
        summary_text = (summary.get('summary_text') or
                       summary.get('summary') or
                       '')
        if summary_text and summary_text != 'N/A':
            if len(summary_text) > 300:
                summary_text = summary_text[:297] + "..."
            results_text += f"<b>📊 Резюме:</b>\n{summary_text}\n\n"

        # Если данных нет или мало, предупреждаем
        if not has_data:
            results_text += "\n⚠️ <i>Анализ не вернул детальных данных.</i>\n"
            results_text += "<i>Возможно, документы были в нестандартном формате.</i>\n\n"

        if html_sent:
            results_text += "<i>💡 Полный HTML отчет отправлен вам в файле</i>"
        elif html_path:
            results_text += "<i>💡 HTML отчет доступен</i>"
        else:
            results_text += "<i>💡 Анализ завершен</i>"

        await callback.message.edit_text(
            results_text,
            parse_mode="HTML",
            reply_markup=get_tender_actions_keyboard(
                tender_index,
                tender_url=tender_url,
                has_analysis=True,
                html_report_path=html_path
            )
        )

    except Exception as e:
        await callback.message.edit_text(
            f"❌ <b>Ошибка при анализе:</b>\n\n"
            f"<code>{str(e)}</code>\n\n"
            f"Попробуйте еще раз позже.",
            reply_markup=get_tender_actions_keyboard(
                tender_index,
                tender_url=tender_url,
                has_analysis=False
            ),
            parse_mode="HTML"
        )


@router.callback_query(SearchStates.viewing_tender_details, F.data.startswith("reanalyze_"))
async def reanalyze_tender(callback: CallbackQuery, state: FSMContext):
    """
    Повторный анализ тендера (для тестирования кэша V2.0).
    """
    await callback.answer()

    # Получаем индекс тендера
    tender_index = int(callback.data.replace("reanalyze_", ""))

    # Получаем результаты поиска из состояния
    data = await state.get_data()
    search_results = data.get('search_results', {})
    results = search_results.get('results', [])

    if tender_index >= len(results):
        await callback.message.answer(
            "❌ Тендер не найден",
            parse_mode="HTML"
        )
        return

    tender_data = results[tender_index]
    tender = tender_data['tender_info']

    # Получаем и нормализуем URL
    tender_url = tender.get('url', '')
    if tender_url and not tender_url.startswith('http'):
        tender_url = f"https://zakupki.gov.ru{tender_url}"

    # Показываем сообщение о начале повторного анализа
    await callback.message.edit_text(
        "🔄 <b>Повторный анализ (тестирование кэша V2.0)...</b>\n\n"
        f"📄 Тендер: {tender.get('number', 'N/A')}\n\n"
        "⏳ Проверяю кэш...\n\n"
        "<i>Если тендер уже анализировался, результат должен загрузиться мгновенно из кэша</i>",
        parse_mode="HTML"
    )

    try:
        # Получаем систему
        system = get_tender_system()

        # Оборачиваем в асинхронный вызов
        loop = asyncio.get_event_loop()

        # Проверяем, были ли уже скачаны документы
        file_paths = None
        tender_dir = results[tender_index].get('tender_dir')

        if tender_dir and results[tender_index].get('documents_downloaded'):
            # Используем уже скачанные документы
            file_paths = [doc['path'] for doc in results[tender_index]['documents_downloaded']]

        # Если документов нет, скачиваем заново
        if not file_paths:
            download_result = await loop.run_in_executor(
                None,
                lambda: system.document_downloader.download_documents(
                    tender_url=tender_url,
                    tender_number=tender.get('number', 'unknown'),
                    doc_types=None
                )
            )

            if download_result['downloaded'] == 0:
                await callback.message.edit_text(
                    "⚠️ <b>Не удалось скачать документы</b>\n\n"
                    f"📄 Тендер: {tender.get('number', 'N/A')}\n\n"
                    "Возможные причины:\n"
                    "• Документы не опубликованы\n"
                    "• Проблема с доступом к сайту\n"
                    "• Неверный URL",
                    reply_markup=get_tender_actions_keyboard(
                        tender_index,
                        tender_url=tender_url,
                        has_analysis=True,
                        html_report_path=results[tender_index].get('html_report_path')
                    ),
                    parse_mode="HTML"
                )
                return

            file_paths = [doc['path'] for doc in download_result.get('files', [])]

        # Импортируем агента для анализа
        from main import TenderAnalysisAgent
        from bot.db import get_database
        import time

        # Создаем агента и анализируем
        agent = TenderAnalysisAgent()

        # Инициализируем БД для кэширования
        agent.db = await get_database()

        # Запоминаем время начала
        start_time = time.time()

        # Анализируем с кэшированием (теперь метод async)
        tender_num = tender.get('number', 'unknown')
        analysis_result = await agent.analyze_tender(
            file_paths,
            tender_number=tender_num,
            use_cache=True  # ВАЖНО: включаем кэш
        )

        # Засекаем время выполнения
        elapsed_time = time.time() - start_time

        # Проверяем, что анализ вернул результаты
        if not analysis_result or not isinstance(analysis_result, dict):
            await callback.message.edit_text(
                "⚠️ <b>Анализ не удался</b>\n\n"
                f"📄 Тендер: {tender.get('number', 'N/A')}\n\n"
                "Возможные причины:\n"
                "• Документы в формате ZIP (не поддерживается)\n"
                "• Не удалось извлечь текст из PDF\n"
                "• Документы защищены паролем\n\n"
                "<i>Попробуйте другой тендер или скачайте документы вручную</i>",
                reply_markup=get_tender_actions_keyboard(
                    tender_index,
                    tender_url=tender_url,
                    has_analysis=True,
                    html_report_path=results[tender_index].get('html_report_path')
                ),
                parse_mode="HTML"
            )
            return

        # Получаем пути к отчетам
        report_paths = analysis_result.get('report_paths', {})
        html_path = report_paths.get('html') if report_paths else None

        # Обновляем данные в состоянии
        results[tender_index]['analysis_result'] = analysis_result
        results[tender_index]['analysis_success'] = True
        results[tender_index]['html_report_path'] = html_path or results[tender_index].get('html_report_path')

        search_results['results'] = results
        await state.update_data(search_results=search_results)

        # Проверяем, был ли использован кэш
        from_cache = analysis_result.get('from_cache', False)

        # Формируем сообщение с результатами
        results_text = "🔄 <b>ПОВТОРНЫЙ АНАЛИЗ ЗАВЕРШЕН</b>\n\n"
        results_text += f"📄 <b>Тендер:</b> {tender.get('number', 'N/A')}\n"
        results_text += f"⏱ <b>Время:</b> {elapsed_time:.2f} сек\n"

        if from_cache:
            results_text += f"💾 <b>Источник:</b> ✅ Кэш (V2.0)\n\n"
            results_text += "<i>🎉 Кэш работает! Анализ загружен мгновенно из базы данных.</i>\n"
        else:
            results_text += f"💾 <b>Источник:</b> 🔄 Новый анализ\n\n"
            results_text += "<i>Тендер проанализирован заново (кэш не найден или устарел).</i>\n"

        results_text += "\n<i>💡 Детальный отчет доступен по кнопке ниже</i>"

        await callback.message.edit_text(
            results_text,
            parse_mode="HTML",
            reply_markup=get_tender_actions_keyboard(
                tender_index,
                tender_url=tender_url,
                has_analysis=True,
                html_report_path=results[tender_index].get('html_report_path')
            )
        )

    except Exception as e:
        logger.error(f"Ошибка при повторном анализе: {e}", exc_info=True)
        await callback.message.edit_text(
            f"❌ <b>Ошибка при повторном анализе:</b>\n\n"
            f"<code>{str(e)}</code>\n\n"
            f"Попробуйте еще раз позже.",
            reply_markup=get_tender_actions_keyboard(
                tender_index,
                tender_url=tender_url,
                has_analysis=True,
                html_report_path=results[tender_index].get('html_report_path')
            ),
            parse_mode="HTML"
        )


@router.callback_query(F.data.startswith("open_report_"))
async def open_html_report(callback: CallbackQuery, state: FSMContext):
    """
    Открыть HTML отчет в браузере.
    """
    await callback.answer()

    # Получаем индекс тендера
    tender_index = int(callback.data.replace("open_report_", ""))

    # Получаем результаты поиска из состояния
    data = await state.get_data()
    search_results = data.get('search_results', {})
    results = search_results.get('results', [])

    if tender_index >= len(results):
        await callback.message.answer(
            "❌ Тендер не найден",
            parse_mode="HTML"
        )
        return

    # Получаем путь к HTML отчету
    html_report_path = results[tender_index].get('html_report_path')

    if not html_report_path:
        await callback.message.answer(
            "❌ HTML отчет не найден",
            parse_mode="HTML"
        )
        return

    try:
        import os

        # Проверяем, что файл существует
        if not os.path.exists(html_report_path):
            await callback.message.answer(
                "❌ Файл HTML отчета не найден",
                parse_mode="HTML"
            )
            return

        # Получаем информацию о тендере для названия файла
        tender = results[tender_index]
        tender_number = tender.get('number', 'unknown')

        # Создаем объект файла для отправки
        document = FSInputFile(html_report_path)

        # Отправляем файл пользователю
        await callback.message.answer_document(
            document=document,
            caption=f"📊 Подробный отчет по тендеру {tender_number}\n\nОткройте файл в браузере для просмотра",
            parse_mode="HTML"
        )

        from aiogram.utils.keyboard import InlineKeyboardBuilder
        from aiogram.types import InlineKeyboardButton

        builder = InlineKeyboardBuilder()
        builder.row(InlineKeyboardButton(
            text="🔍 Новый поиск",
            callback_data="main_menu"
        ))
        builder.row(InlineKeyboardButton(
            text="🏠 Главное меню",
            callback_data="main_menu"
        ))

        await callback.message.answer(
            "✅ HTML отчет отправлен! Откройте файл для просмотра подробной информации.",
            reply_markup=builder.as_markup(),
            parse_mode="HTML"
        )

    except Exception as e:
        await callback.message.answer(
            f"❌ Не удалось отправить HTML отчет:\n\n<code>{str(e)}</code>",
            parse_mode="HTML"
        )


@router.callback_query(F.data == "main_menu")
async def return_to_main_menu(callback: CallbackQuery, state: FSMContext):
    """
    Вернуться в главное меню.
    """
    await callback.answer()
    await state.clear()

    await callback.message.edit_text(
        "🏠 Возвращаемся в главное меню...",
        parse_mode="HTML"
    )

    await callback.message.answer(
        "Выберите действие:",
        reply_markup=get_main_menu_keyboard(),
        parse_mode="HTML"
    )


@router.callback_query(F.data == "new_search")
async def start_new_search(callback: CallbackQuery, state: FSMContext):
    """
    Начать новый поиск - очищает состояние и возвращает к вводу запроса.
    """
    await callback.answer("🔄 Начинаем новый поиск...")
    await state.clear()

    await callback.message.edit_text(
        "🔍 <b>Новый поиск тендеров</b>\n\n"
        "Введите поисковый запрос (название товаров/услуг):\n\n"
        "<i>Например: «компьютеры и оргтехника», «строительные материалы», «канцелярские товары»</i>",
        parse_mode="HTML",
        reply_markup=get_inline_cancel_keyboard()
    )

    await state.set_state(SearchStates.waiting_for_query)


@router.callback_query(SearchStates.viewing_results, F.data == "batch_analyze_all")
async def batch_analyze_all_tenders(callback: CallbackQuery, state: FSMContext):
    """
    Пакетный анализ всех найденных тендеров.
    """
    await callback.answer()

    # Получаем результаты поиска из состояния
    data = await state.get_data()
    search_results = data.get('search_results', {})
    results = search_results.get('results', [])

    if not results:
        await callback.message.answer(
            "❌ Нет тендеров для анализа",
            parse_mode="HTML"
        )
        return

    total_tenders = len(results)

    # Показываем начальное сообщение
    progress_message = await callback.message.edit_text(
        f"📦 <b>Пакетный анализ запущен</b>\n\n"
        f"📊 Тендеров: {total_tenders}\n"
        f"⚡ Параллельность: 3x\n\n"
        f"<b>Прогресс:</b>\n"
        f"{'█' * 0}{'░' * 10} 0% (0/{total_tenders})\n"
        f"⏱️ Начинаем...\n\n"
        f"💚 Из кэша: 0\n"
        f"🔄 Новых: 0",
        parse_mode="HTML"
    )

    # Получаем систему
    system = get_tender_system()
    loop = asyncio.get_event_loop()

    # Результаты анализа
    analyzed_results = []
    from_cache = 0
    new_analysis = 0

    # Импортируем агента для анализа
    from main import TenderAnalysisAgent
    from bot.db import get_database

    # Создаем агента один раз для всех тендеров
    agent = TenderAnalysisAgent()
    agent.db = await get_database()

    # Анализируем тендеры с прогресс-баром
    for i, tender_data in enumerate(results):
        tender = tender_data['tender_info']
        tender_url = tender.get('url', '')
        if tender_url and not tender_url.startswith('http'):
            tender_url = f"https://zakupki.gov.ru{tender_url}"

        # Обновляем прогресс
        progress = int((i / total_tenders) * 10)
        percent = int((i / total_tenders) * 100)

        await progress_message.edit_text(
            f"📦 <b>Пакетный анализ</b>\n\n"
            f"📊 Тендеров: {total_tenders}\n"
            f"⚡ Параллельность: 3x\n\n"
            f"<b>Прогресс:</b>\n"
            f"{'█' * progress}{'░' * (10 - progress)} {percent}% ({i}/{total_tenders})\n"
            f"⏱️ Анализируем тендер #{i+1}...\n\n"
            f"💚 Из кэша: {from_cache}\n"
            f"🔄 Новых: {new_analysis}",
            parse_mode="HTML"
        )

        try:
            # Шаг 1: Скачиваем документы
            download_result = await loop.run_in_executor(
                None,
                lambda url=tender_url, num=tender.get('number', 'unknown'): system.document_downloader.download_documents(
                    tender_url=url,
                    tender_number=num,
                    doc_types=None
                )
            )

            if download_result['downloaded'] == 0:
                analyzed_results.append({
                    'tender': tender,
                    'analysis': None,
                    'error': 'Не удалось скачать документы',
                    'index': i
                })
                continue

            # Шаг 2: Анализируем документы
            file_paths = [doc['path'] for doc in download_result.get('files', [])]
            tender_num = tender.get('number', 'unknown')

            analysis_result = await agent.analyze_tender(
                file_paths,
                tender_number=tender_num,
                use_cache=True
            )

            # Проверяем был ли взят из кэша
            if analysis_result.get('from_cache'):
                from_cache += 1
            else:
                new_analysis += 1

            analyzed_results.append({
                'tender': tender,
                'analysis': analysis_result,
                'index': i
            })

        except Exception as e:
            print(f"Ошибка анализа тендера {i}: {e}")
            import traceback
            traceback.print_exc()
            analyzed_results.append({
                'tender': tender,
                'analysis': None,
                'error': str(e),
                'index': i
            })

    # Финальный прогресс
    await progress_message.edit_text(
        f"✅ <b>Пакетный анализ завершен!</b>\n\n"
        f"📊 Всего тендеров: {total_tenders}\n"
        f"✅ Успешно: {len([r for r in analyzed_results if r.get('analysis')])}\n"
        f"❌ Ошибок: {len([r for r in analyzed_results if r.get('error')])}\n\n"
        f"💚 Из кэша: {from_cache}\n"
        f"🔄 Новых: {new_analysis}\n\n"
        f"<i>Формирую результаты...</i>",
        parse_mode="HTML"
    )

    # Сохраняем результаты в state
    await state.update_data(batch_analysis_results=analyzed_results)

    # Показываем результаты
    await show_batch_results(callback.message, state, analyzed_results)


async def show_batch_results(message, state: FSMContext, results: list):
    """
    Показывает результаты пакетного анализа.
    """
    if not results:
        await message.edit_text(
            "❌ Нет результатов для отображения",
            parse_mode="HTML"
        )
        return

    results_text = "📋 <b>РЕЗУЛЬТАТЫ ПАКЕТНОГО АНАЛИЗА</b>\n\n"

    for result in results:
        tender = result['tender']
        analysis = result.get('analysis')
        error = result.get('error')
        index = result['index']

        # Извлекаем данные
        number = tender.get('number', 'N/A')
        name = tender.get('name', 'Без названия')
        if len(name) > 70:
            name = name[:67] + "..."

        results_text += f"{index + 1}. <b>№ {number}</b>\n"
        results_text += f"   📦 {name}\n"

        if error:
            results_text += f"   ❌ Ошибка анализа\n\n"
        elif analysis:
            tender_info = analysis.get('tender_info', {})
            gaps = analysis.get('gaps', [])

            nmck = tender_info.get('nmck', 0)
            prepayment = tender_info.get('prepayment_percent', 0)

            if nmck:
                results_text += f"   💰 НМЦК: {nmck:,.0f} руб.\n"
            if prepayment:
                results_text += f"   💳 Аванс: {prepayment}%\n"

            critical_gaps = sum(1 for g in gaps if isinstance(g, dict) and g.get('severity') == 'CRITICAL')
            if critical_gaps > 0:
                results_text += f"   ⚠️ Критичных пробелов: {critical_gaps}\n"

            results_text += "\n"
        else:
            results_text += f"   ⏳ Не проанализирован\n\n"

    results_text += "\n<i>💡 Нажмите на номер тендера для подробной информации</i>"

    # Создаем клавиатуру с кнопками для каждого тендера
    from aiogram.utils.keyboard import InlineKeyboardBuilder
    from aiogram.types import InlineKeyboardButton

    builder = InlineKeyboardBuilder()

    for result in results:
        index = result['index']
        tender = result['tender']
        number = tender.get('number', 'N/A')[:15]

        builder.row(
            InlineKeyboardButton(
                text=f"📄 {number}",
                callback_data=f"details_{index}"
            )
        )

    # Кнопка экспорта в Excel
    builder.row(
        InlineKeyboardButton(text="📊 Экспорт в Excel", callback_data="export_excel")
    )

    builder.row(
        InlineKeyboardButton(text="🔄 Новый поиск", callback_data="new_search"),
        InlineKeyboardButton(text="🏠 Главное меню", callback_data="main_menu")
    )

    await message.edit_text(
        results_text,
        parse_mode="HTML",
        reply_markup=builder.as_markup()
    )


@router.callback_query(SearchStates.viewing_results, F.data == "export_excel")
async def export_to_excel(callback: CallbackQuery, state: FSMContext):
    """
    Экспорт результатов пакетного анализа в Excel.
    """
    await callback.answer("📊 Формирую Excel файл...")

    # Получаем результаты из state
    data = await state.get_data()
    batch_results = data.get('batch_analysis_results', [])

    if not batch_results:
        await callback.message.answer(
            "❌ Нет данных для экспорта. Сначала выполните пакетный анализ.",
            parse_mode="HTML"
        )
        return

    try:
        # Создаем Excel файл
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from datetime import datetime
        import os

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Анализ тендеров"

        # Заголовок
        headers = [
            "№ п/п",
            "Номер тендера",
            "Объект закупки",
            "НМЦК (руб.)",
            "Аванс (%)",
            "Обеспечение заявки (руб.)",
            "Обеспечение контракта (руб.)",
            "Критичные пробелы",
            "Важные пробелы",
            "Средние пробелы",
            "Низкие пробелы",
            "Срок подачи",
            "Статус анализа"
        ]

        # Стиль заголовка
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Записываем данные
        for row_num, result in enumerate(batch_results, 2):
            tender = result['tender']
            analysis = result.get('analysis')
            error = result.get('error')

            # № п/п
            ws.cell(row=row_num, column=1).value = row_num - 1

            # Номер тендера
            ws.cell(row=row_num, column=2).value = tender.get('number', 'N/A')

            # Объект закупки
            ws.cell(row=row_num, column=3).value = tender.get('name', 'Без названия')

            if error:
                ws.cell(row=row_num, column=13).value = f"Ошибка: {error}"
                continue

            if analysis:
                tender_info = analysis.get('tender_info', {})
                financial = analysis.get('financial_analysis', {})
                gaps = analysis.get('gaps', [])

                # НМЦК
                nmck = tender_info.get('nmck', 0)
                ws.cell(row=row_num, column=4).value = nmck
                ws.cell(row=row_num, column=4).number_format = '#,##0'

                # Аванс
                prepayment = tender_info.get('prepayment_percent', 0)
                ws.cell(row=row_num, column=5).value = prepayment

                # Обеспечение заявки
                guarantees = financial.get('guarantees', {})
                app_guarantee = guarantees.get('application_guarantee', 0)
                ws.cell(row=row_num, column=6).value = app_guarantee
                ws.cell(row=row_num, column=6).number_format = '#,##0'

                # Обеспечение контракта
                contract_guarantee = guarantees.get('contract_guarantee', 0)
                ws.cell(row=row_num, column=7).value = contract_guarantee
                ws.cell(row=row_num, column=7).number_format = '#,##0'

                # Пробелы по критичности
                critical = sum(1 for g in gaps if isinstance(g, dict) and g.get('severity') == 'CRITICAL')
                high = sum(1 for g in gaps if isinstance(g, dict) and g.get('severity') == 'HIGH')
                medium = sum(1 for g in gaps if isinstance(g, dict) and g.get('severity') == 'MEDIUM')
                low = sum(1 for g in gaps if isinstance(g, dict) and g.get('severity') == 'LOW')

                ws.cell(row=row_num, column=8).value = critical
                ws.cell(row=row_num, column=9).value = high
                ws.cell(row=row_num, column=10).value = medium
                ws.cell(row=row_num, column=11).value = low

                # Срок подачи
                deadline = tender_info.get('deadline_submission', 'N/A')
                ws.cell(row=row_num, column=12).value = deadline

                # Статус
                ws.cell(row=row_num, column=13).value = "✅ Проанализирован"
            else:
                ws.cell(row=row_num, column=13).value = "⏳ Не проанализирован"

        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Сохраняем файл
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"tender_analysis_{timestamp}.xlsx"
        filepath = f"/tmp/{filename}"

        wb.save(filepath)

        # Отправляем файл пользователю
        from aiogram.types import FSInputFile

        document = FSInputFile(filepath, filename=filename)

        await callback.message.answer_document(
            document=document,
            caption=f"📊 <b>Excel отчет готов</b>\n\n"
                   f"📋 Тендеров: {len(batch_results)}\n"
                   f"✅ Проанализировано: {len([r for r in batch_results if r.get('analysis')])}\n"
                   f"❌ Ошибок: {len([r for r in batch_results if r.get('error')])}\n\n"
                   f"<i>Файл содержит детальную информацию по каждому тендеру</i>",
            parse_mode="HTML"
        )

        # Удаляем временный файл
        try:
            os.remove(filepath)
        except:
            pass

        await callback.answer("✅ Excel файл отправлен")

    except Exception as e:
        print(f"Ошибка экспорта в Excel: {e}")
        await callback.answer("❌ Ошибка при создании Excel файла", show_alert=True)
