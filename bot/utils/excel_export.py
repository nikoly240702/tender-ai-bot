"""
Excel Export Utility.

Экспорт тендеров в формат Excel (.xlsx).
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Путь для сохранения отчётов
REPORTS_DIR = Path(__file__).parent.parent.parent / "output" / "excel_reports"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def generate_tenders_excel(
    tenders: List[Dict[str, Any]],
    user_id: int,
    title: str = "Мои тендеры",
    filter_name: Optional[str] = None
) -> Path:
    """
    Генерирует Excel файл с тендерами.

    Args:
        tenders: Список тендеров
        user_id: ID пользователя
        title: Заголовок отчёта
        filter_name: Название фильтра (если есть)

    Returns:
        Path: Путь к созданному файлу
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Тендеры"

    # Стили
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(vertical="top", wrap_text=True)
    border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )

    # Заголовки колонок
    headers = [
        ("№", 5),
        ("Название", 50),
        ("Цена (НМЦ)", 15),
        ("Заказчик", 35),
        ("Регион", 20),
        ("Дедлайн", 12),
        ("Закон", 8),
        ("Статус", 12),
        ("Ссылка", 40),
    ]

    # Записываем заголовки
    for col, (header_name, width) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Альтернативные цвета строк
    alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    # Записываем данные
    for row_idx, tender in enumerate(tenders, 2):
        # Извлекаем данные
        name = tender.get('name') or tender.get('title', 'Без названия')
        price = tender.get('price') or tender.get('max_price', 0)
        customer = tender.get('customer') or tender.get('organization', '-')
        region = tender.get('region') or tender.get('delivery_region', '-')
        deadline = tender.get('deadline') or tender.get('end_date', '-')
        law = tender.get('law_type') or tender.get('purchase_type', '-')
        status = tender.get('status', 'Активен')
        link = tender.get('link') or tender.get('url', '')

        # Форматируем цену
        if isinstance(price, (int, float)) and price > 0:
            price_str = f"{price:,.0f} ₽".replace(",", " ")
        else:
            price_str = "-"

        # Форматируем дедлайн
        if isinstance(deadline, datetime):
            deadline_str = deadline.strftime("%d.%m.%Y")
        elif isinstance(deadline, str) and deadline != '-':
            deadline_str = deadline[:10] if len(deadline) > 10 else deadline
        else:
            deadline_str = "-"

        # Формируем ссылку на zakupki.gov.ru
        if not link and tender.get('number'):
            link = f"https://zakupki.gov.ru/epz/order/notice/ea20/view/common-info.html?regNumber={tender['number']}"

        row_data = [
            row_idx - 1,  # №
            name[:200],  # Обрезаем длинные названия
            price_str,
            customer[:100] if customer else "-",
            region[:50] if region else "-",
            deadline_str,
            law[:10] if law else "-",
            status,
            link,
        ]

        # Записываем строку
        row_fill = alt_fill if row_idx % 2 == 0 else None
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = border
            if row_fill:
                cell.fill = row_fill

    # Делаем ссылки кликабельными
    for row_idx in range(2, len(tenders) + 2):
        cell = ws.cell(row=row_idx, column=9)  # Колонка ссылки
        if cell.value and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.font = Font(color="0066CC", underline="single")

    # Добавляем итоговую строку
    total_row = len(tenders) + 3
    ws.cell(row=total_row, column=1, value="Итого:").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=f"{len(tenders)} тендеров").font = Font(bold=True)

    # Суммируем цены
    total_price = sum(
        t.get('price') or t.get('max_price', 0)
        for t in tenders
        if isinstance(t.get('price') or t.get('max_price'), (int, float))
    )
    if total_price > 0:
        ws.cell(row=total_row, column=3, value=f"{total_price:,.0f} ₽".replace(",", " ")).font = Font(bold=True)

    # Генерируем имя файла
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_title = "".join(c for c in title if c.isalnum() or c in " _-")[:30]
    filename = f"tenders_{safe_title}_{user_id}_{timestamp}.xlsx"
    filepath = REPORTS_DIR / filename

    # Сохраняем
    wb.save(filepath)
    logger.info(f"Excel отчёт создан: {filepath} ({len(tenders)} тендеров)")

    return filepath


async def generate_tenders_excel_async(
    tenders: List[Dict[str, Any]],
    user_id: int,
    title: str = "Мои тендеры",
    filter_name: Optional[str] = None
) -> Path:
    """
    Асинхронная обёртка для генерации Excel.
    """
    import asyncio
    return await asyncio.to_thread(
        generate_tenders_excel,
        tenders,
        user_id,
        title,
        filter_name
    )
